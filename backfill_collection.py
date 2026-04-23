#!/usr/bin/env python3
"""
Backfill `collections` rows from Excel trackers + dp_sheet + PostgreSQL.

Dry-run (default, no DB writes):
    python3 backfill_collection.py

Execute updates:
    python3 backfill_collection.py --execute

After each run, a CSV is written under the data directory:
    cannot_update_collections_<timestamp>.csv
  listing every export collection_id that was not updated (or would not be in dry-run), with loan_id,
  Application Number (from LOS), and reason.

Expected files (same directory as this script unless BACKFILL_DATA_DIR is set):
  - db_sheet.xlsx            Sheet1:  loan_id (source list); often the same export holds DP fields
  - LOS _Data.xlsx           Sheet4:  application_id (same value as loan_id in DB), Application Number
  - COMBINE-EMI-FOUR-ZONES.xlsx  Sheet1:  grouped EMI blocks:  EMI - k  then  Paid On, Cash Amount,
        Online Amount, Total Amount, Comments  (repeat for each k)
  - dp_sheet.xlsx            Sheet1:  optional separate file — if absent, CONFIG points to db_sheet.xlsx

openpyxl must use read_only=False here: in read_only mode worksheet dimensions are unknown until
iteration, so max_column stays None and header discovery scans only column A (silent failure).

Validation: only COMBINE Tenure must equal DP `tenure`. `no_of_actual_emis` (count of EMI Paid On
dates in COMBINE) is logged for audit and debug, not used to reject a loan.

DP export (`db_sheet`) lists only **pending** collection rows — not every EMI 1..T. Updates run only
for EMI numbers that appear both as **paid** in COMBINE (`EMI - k` / Paid On) and as a row in the
DP export with matching `emi_installment_no`.

There are no alternate column names or sheet fallbacks: missing names halt with ERROR + HALT.
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg2 import sql

# -----------------------------------------------------------------------------
# Config — exact labels (no fallbacks). Edit here if your export names change.
# -----------------------------------------------------------------------------

DATA_DIR_ENV = "BACKFILL_DATA_DIR"

FILES = {
    "db": ("db_sheet.xlsx", "Sheet1"),
    "los": ("LOS _Data.xlsx", "Sheet4"),
    "combine": ("Merged-EMI'S.xlsx", "Sheet8"),
    # Row-level collection export (tenure, monthly_payable, collection_id, …). Use db_sheet if no separate file.
    "dp": ("db_sheet.xlsx", "Sheet1"),
}

# db_sheet
COL_DB_LOAN_ID = "loan_id"

# LOS _Data.xlsx — export uses application_id (joins to collections.loan_id / application_id)
COL_LOS_LOAN_ID = "application_id"
COL_LOS_APP_NUMBER = "Application Number"
COL_LOS_CUSTOMER_ID = "customer_id"

# COMBINE tracker
COL_CT_APP = "Application No."
COL_CT_TENURE = "Tenure"
COL_CT_EMI = "EMI"
COL_CT_CLOSE_TYPE = "CLOSE_TYPE"

# DP export — exact header strings (evaluated on sample: monthly_payble typo in export)
COL_DP_LOAN_ID = "loan_id"
COL_DP_TENURE = "tenure"
COL_DP_MONTHLY = "monthly_payble"
COL_DP_DUE_AMOUNT = "due_amount"
COL_DP_CENTER_ID = "center_id"
COL_DP_COLLECTION_ID = "collection_id"
COL_DP_EMI_NO = "emi_installment_no"
COL_DP_STATUS = "status"
COL_DP_SUBTYPE = "collection_subtype"
# Optional: if missing from header row, part-payment rule uses tracker Total Amount only for threshold checks
COL_DP_P1 = "p1"
COL_DP_P2 = "p2"

STATUS_PENDING = "PENDING"
STATUS_DONE = "DONE"
# Amount matching mode:
# - "exact": COMBINE Total Amount must equal due_amount exactly (no fine/discount path)
# - "unlimited": always allow difference (fine for extra, discount for shortfall)
# - "capped": allow difference only when <= MAX_ALLOWED_AMOUNT_DIFF
AMOUNT_MATCH_MODE = "exact"
# Amount-difference control for fine/discount calculation.
# - None: no limit (always allow full extra/shortfall as fine/discount)
# - Decimal("200"): allow only when absolute difference <= 200
# - Decimal("1000"), Decimal("5000"), etc. for future policy changes
MAX_ALLOWED_AMOUNT_DIFF: Decimal | None = None

CENTER_MANAGER_TO_CENTER_IDS = {
    1072: [2],  # PANIPAT
    1060: [1, 3, 7, 8, 15],  # SONIPAT, KARNAL, CHANDIGARH, MEERUT, BIJNOR
    1084: [4, 11],  # LUCKNOW, LUCKNOW_TDP
    1071: [5],  # PRAYAGRAJ
    427: [9],  # SAHARANPUR
    1489: [10],  # MATHURA
    2368: [13],  # BAREILLY
    2017: [12],  # VARANASI
    3326: [14],  # HAPUR
    3583: [6],  # JAIPUR
    3819: [16],  # AGRA
    3911: [17],  # MUZAFFARNAGAR
    4421: [18],  # GHAZIABAD
    4428: [19],  # NOIDA
}

# DB
ENV_DB_HOST = "DB_POSTGRES_URL"  # historical name in prompt; used as host
ENV_DB_NAME = "DB_POSTGRES_DBNAME"
ENV_DB_USER = "DB_POSTGRES_USERNAME"
ENV_DB_PASS = "DB_POSTGRES_PASS"
ENV_DB_PORT = "DB_POSTGRES_PORT"

TABLE_COLLECTIONS = "collections"

# COMBINE: leading EMI group title cell, e.g. "EMI - 1" (spaces around dash)
RE_EMI_GROUP_LABEL = re.compile(r"^EMI\s*-\s*(\d+)\s*$", re.IGNORECASE)
SUB_PAID_ON = "Paid On"
SUB_CASH_AMOUNT = "Cash Amount"
SUB_ONLINE_AMOUNT = "Online Amount"
SUB_TOTAL_AMOUNT = "Total Amount"
SUB_COMMENTS = "Comments"
SKIP_CLOSE_TYPES = {"recovered", "closed", "return"}


def is_part_subtype(raw: str | None) -> bool:
    if raw is None or cell_str(raw) == "":
        return False
    s = cell_str(raw).lower().replace(" ", "").replace("-", "_")
    if s in ("part_1", "part_2"):
        return True
    if s in ("part1", "part2"):
        return True
    return False


def dp_row_subtype_bucket(sub: str | None) -> str:
    """
    Rows under the same (loan_id, emi_installment_no) must be unique per bucket.
    PART1 and PART2 are separate buckets so two rows (part 1 + part 2) are allowed;
    any other duplicate bucket is a HALT at load time.
    """
    if sub is None or cell_str(sub).strip() == "":
        return "REG"
    t = cell_str(sub).upper().replace(" ", "").replace("-", "_")
    if t in ("PART1", "PART_1", "P1"):
        return "PART1"
    if t in ("PART2", "PART_2", "P2"):
        return "PART2"
    if "PART" in t:
        if "2" in t:
            return "PART2"
        return "PART1"
    return f"REG:{t}"


@dataclass(frozen=True)
class DpEmiRow:
    collection_id: int
    emi_installment_no: int
    tenure: int
    monthly_payable: Decimal
    due_amount: Decimal
    center_id: int
    status: str
    collection_sub_type: str | None
    p1: Decimal | None
    p2: Decimal | None


def setup_logging(log_prefix: str = "backfill") -> logging.Logger:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = Path.cwd() / f"{log_prefix}_{ts}.log"
    logger = logging.getLogger("backfill_collection")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    ch.setLevel(logging.INFO)
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.debug("Log file: %s", log_path.resolve())
    return logger


def load_env_file(logger: logging.Logger, path: Path) -> None:
    """
    Load KEY=VALUE pairs from .env into process environment.
    Existing environment variables are overwritten by .env values.
    """
    if not path.is_file():
        logger.warning("Checkpoint: .env not found at %s (continuing with current environment)", path)
        return
    loaded = 0
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
            value = value[1:-1]
        os.environ[key] = value
        loaded += 1
    logger.info("Checkpoint: loaded %s env vars from %s", loaded, path.name)


def _merge_cannot_update_reason(existing: str, new: str) -> str:
    if new in existing or existing.endswith(new):
        return existing
    return f"{existing}; {new}"


def note_cannot_update_row(
    blocked: dict[int, dict[str, str]],
    los: dict[int, str],
    *,
    collection_id: int,
    loan_id: int,
    reason: str,
) -> None:
    app = cell_str(los.get(loan_id, "") or "")
    prev = blocked.get(collection_id)
    if prev is None:
        blocked[collection_id] = {
            "loan_id": str(loan_id),
            "application_no": app,
            "reason": reason,
        }
        return
    prev["reason"] = _merge_cannot_update_reason(prev["reason"], reason)
    if app and not prev.get("application_no"):
        prev["application_no"] = app


def note_all_pending_collections_for_loan(
    blocked: dict[int, dict[str, str]],
    los: dict[int, str],
    dp_by_loan: dict[int, dict[int, list[DpEmiRow]]],
    loan_id: int,
    reason: str,
) -> None:
    em = dp_by_loan.get(loan_id)
    if not em:
        return
    for rows in em.values():
        for r in rows:
            if cell_str(r.status).upper() == STATUS_PENDING:
                note_cannot_update_row(
                    blocked,
                    los,
                    collection_id=r.collection_id,
                    loan_id=loan_id,
                    reason=reason,
                )


REASON_DESCRIPTIONS = {
    "loan_not_in_LOS": "Loan ID exists in dp_sheet but is not present in LOS mapping.",
    "no_COMBINE_for_loan": "No matching loan row found in COMBINE tracker for this loan.",
    "no_db_sheet_rows_for_loan": "Loan is present in loan list but no dp_sheet rows were found for it.",
    "combine_close_type_blocked": "Loan skipped because COMBINE CLOSE_TYPE is Recovered/Closed/Return.",
    "dp_tenure_missing": "DP tenure value for the loan is missing or invalid.",
    "dp_emi_outside_combine_tenure": "EMI in dp_sheet is outside COMBINE tenure range.",
    "combine_paid_on_missing_or_invalid": "COMBINE Paid On for this EMI is blank/invalid, so not eligible.",
    "combine_total_amount_not_usable": "COMBINE Total Amount is blank/NA/non-numeric for this EMI.",
    "part_payment_not_handled": "Part payment subtype is intentionally skipped by policy.",
    "center_id_manager_mapping_missing": "center_id has no configured user mapping for created_by/collector_id.",
    "amount_mismatch_exact_required": "Exact match mode enabled; COMBINE Total Amount != due_amount.",
    "extra_amount_exceeds_limit": "COMBINE Total Amount exceeds due_amount beyond configured cap.",
    "shortfall_exceeds_limit": "COMBINE Total Amount is below due_amount beyond configured cap.",
    "DB_collection_row_not_found_or_inactive": "collection_id not found in active DB records.",
    "DB_status_not_PENDING": "DB status for collection_id is not PENDING.",
    "UPDATE_rowcount_not_1": "DB update affected 0/multiple rows instead of exactly 1.",
    "no_eligible_path_COMBINE_paid_on_missing_or_invalid_for_this_emi": (
        "No eligible update path found (typically COMBINE Paid On missing/invalid for this EMI)."
    ),
}


def describe_reason(reason: str) -> str:
    parts = [p.strip() for p in cell_str(reason).split(";") if p.strip()]
    if not parts:
        return ""
    descs: list[str] = []
    for p in parts:
        descs.append(REASON_DESCRIPTIONS.get(p, f"Reason '{p}' occurred."))
    return " | ".join(descs)


def write_cannot_update_collections_csv(
    path: Path,
    blocked: dict[int, dict[str, str]],
    would_update_ids: set[int],
    logger: logging.Logger,
) -> None:
    """Write collection_id, loan_id, application_no, reason for rows not in would_update_ids."""
    rows_out: list[tuple[int, str, str, str, str]] = []
    for cid, rec in sorted(blocked.items(), key=lambda x: x[0]):
        if cid in would_update_ids:
            continue
        reason = rec.get("reason", "")
        rows_out.append(
            (
                cid,
                rec.get("loan_id", ""),
                rec.get("application_no", ""),
                reason,
                describe_reason(reason),
            )
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["collection_id", "loan_id", "application_no", "reason", "reason desc"])
        w.writerows(rows_out)
    logger.info(
        "Checkpoint: cannot-update report %s (%s collection_id rows)",
        path.resolve(),
        len(rows_out),
    )


def write_all_collection_ids_csv(
    path: Path,
    dp_by_loan: dict[int, dict[int, list[DpEmiRow]]],
    los: dict[int, str],
    blocked: dict[int, dict[str, str]],
    would_update_ids: set[int],
    logger: logging.Logger,
) -> None:
    """Write one row for every collection_id from dp export with outcome and reason."""
    rows_out: list[tuple[int, str, str, str, str]] = []
    for loan_id in sorted(dp_by_loan.keys()):
        app_no = cell_str(los.get(loan_id, "") or "")
        for emi_no in sorted(dp_by_loan[loan_id].keys()):
            for r in dp_by_loan[loan_id][emi_no]:
                status_in_dp = cell_str(r.status).upper()
                cid = r.collection_id
                if cid in would_update_ids:
                    outcome = "would_update_or_updated"
                    reason = ""
                else:
                    outcome = "not_updated"
                    if status_in_dp != STATUS_PENDING:
                        reason = "non_pending_in_dp_export"
                    else:
                        rec = blocked.get(cid)
                        reason = rec.get("reason", "") if rec else ""
                rows_out.append((cid, str(loan_id), app_no, outcome, reason))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["collection_id", "loan_id", "application_no", "outcome", "reason"])
        w.writerows(rows_out)
    logger.info(
        "Checkpoint: all-collection report %s (%s rows)",
        path.resolve(),
        len(rows_out),
    )


def write_updated_collections_xlsx(
    path: Path,
    rows: list[tuple[int, int, str, int, str, str]],
    logger: logging.Logger,
) -> None:
    """Write updated/would-update collection rows to XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "updated_collections"
    ws.append(
        [
            "collection_id",
            "loan_id",
            "application_no",
            "emi_installment_no",
            "collection_subtype",
            "due_amount",
        ]
    )
    for row in rows:
        ws.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    logger.info("Checkpoint: updated-collections xlsx %s (%s rows)", path.resolve(), len(rows))


def write_amount_adjustments_xlsx(
    path: Path,
    rows: list[tuple[int, int, str, int, str, str, str, str, str]],
    logger: logging.Logger,
) -> None:
    """Write rows where fine_amount or discount_amount is applied."""
    wb = Workbook()
    ws = wb.active
    ws.title = "amount_adjustments"
    ws.append(
        [
            "collection_id",
            "loan_id",
            "application_no",
            "emi_installment_no",
            "collection_subtype",
            "combine_total_amount",
            "due_amount",
            "fine_amount",
            "discount_amount",
        ]
    )
    for row in rows:
        ws.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    logger.info("Checkpoint: amount-adjustments xlsx %s (%s rows)", path.resolve(), len(rows))


def paid_on_to_iso_timestamp(value: Any) -> str:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)):
        dt = from_excel(float(value))
        if isinstance(dt, date) and not isinstance(dt, datetime):
            dt = datetime.combine(dt, time.min)
    else:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def write_phase2_collection_trans_csv(
    path: Path,
    trans_rows: list[tuple[Any, ...]],
    logger: logging.Logger,
) -> None:
    headers = [
        "trans_id",
        "is_aggr_trans",
        "org_id",
        "center_id",
        "collection_id",
        "is_active",
        "transaction_cd",
        "trans_type",
        "trans_subtype",
        "mode",
        "amount",
        "recorded_by",
        "scrap_type_id",
        "recon_status",
        "recon_by",
        "recon_at",
        "recon_comment",
        "is_settled",
        "from_user",
        "from_customer",
        "to_user",
        "created_by",
        "created_at",
        "trans_comments",
        "parent_trans_id",
        "request_id",
        "deposit_trans_ids",
        "actual_created_at",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(trans_rows)
    logger.info(
        "PHASE 2: collection_trans csv generated at %s (%s rows)",
        path.resolve(),
        len(trans_rows),
    )


def write_phase2_collection_comments_csv(
    path: Path,
    comment_rows: list[tuple[Any, ...]],
    logger: logging.Logger,
) -> None:
    headers = [
        "collection_comment_id",
        "org_id",
        "collection_id",
        "is_called",
        "comment",
        "is_latest",
        "created_by",
        "created_at",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(comment_rows)
    logger.info(
        "PHASE 2: collection_comments csv generated at %s (%s rows)",
        path.resolve(),
        len(comment_rows),
    )


def write_part1_tracker_analysis_xlsx(
    path: Path,
    dp_by_loan: dict[int, dict[int, list[DpEmiRow]]],
    los: dict[int, str],
    combine_by_app: dict[str, dict[str, Any]],
    logger: logging.Logger,
) -> None:
    """
    Build PART1-only tracker comparison output:
      - Pending_PART1 sheet
      - Done_PART1 sheet
    is_paid_in_tracker=True only when:
      - COMBINE Paid On for EMI-k is valid, and
      - COMBINE Total Amount exactly equals PART1 due_amount.
    """
    pending_rows: list[tuple[int, int, str, str, str, bool, str, str, str, str, bool, bool]] = []
    done_rows: list[tuple[int, int, str, str, str, bool, str, str, str, str, bool, bool]] = []
    missing_or_unmatched: Counter[str] = Counter()

    for loan_id in sorted(dp_by_loan.keys()):
        app = los.get(loan_id)
        app_key = normalize_app_number_key(app) if app else ""
        cmb = combine_by_app.get(app_key) if app_key else None
        for emi_no in sorted(dp_by_loan[loan_id].keys()):
            for r in dp_by_loan[loan_id][emi_no]:
                if dp_row_subtype_bucket(r.collection_sub_type) != "PART1":
                    continue
                status_up = cell_str(r.status).upper()
                is_paid_in_tracker = False
                is_amount_matched = False
                paid_on_valid = False
                tracker_amount = ""
                due_amount = str(r.due_amount)
                monthly_payble = str(r.monthly_payable)
                amount_difference = ""
                tracker_amount_non_zero_numeric = False
                reason = ""
                if cmb is None:
                    reason = "missing_combine_for_loan"
                else:
                    paid_on_raw = cmb["paid_on"].get(emi_no)
                    paid_on_valid = is_valid_paid_on(paid_on_raw)
                    tot_raw = cmb["total"].get(emi_no)
                    tot_dec = parse_combine_total_amount(tot_raw)
                    if tot_dec is None:
                        reason = "missing_or_invalid_combine_total_amount"
                    else:
                        tracker_amount = str(tot_dec)
                        tracker_amount_non_zero_numeric = tot_dec != Decimal("0")
                        diff = tot_dec - r.due_amount
                        amount_difference = str(diff)
                        is_amount_matched = tot_dec == r.due_amount
                        is_paid_in_tracker = paid_on_valid and tracker_amount_non_zero_numeric
                        if not paid_on_valid:
                            reason = "paid_on_missing_or_invalid"
                        elif not is_amount_matched:
                            reason = "exact_amount_mismatch"

                if reason:
                    missing_or_unmatched[reason] += 1

                out = (
                    r.collection_id,
                    loan_id,
                    cell_str(app),
                    cell_str(r.collection_sub_type),
                    status_up,
                    is_amount_matched,
                    tracker_amount,
                    due_amount,
                    monthly_payble,
                    amount_difference,
                    paid_on_valid,
                    is_paid_in_tracker,
                )
                if status_up == STATUS_PENDING:
                    pending_rows.append(out)
                elif status_up == STATUS_DONE:
                    done_rows.append(out)

    wb = Workbook()
    ws_pending = wb.active
    ws_pending.title = "Pending_PART1"
    headers = [
        "collection_id",
        "loan_id",
        "application_no",
        "collection_sub_type",
        "db_status",
        "is_amount_matched",
        "tracker_amount",
        "collection_due_amount",
        "monthly_payble",
        "amount_difference",
        "paid_on_valid",
        "is_paid_in_tracker",
    ]
    ws_pending.append(headers)
    for row in pending_rows:
        ws_pending.append(list(row))

    ws_done = wb.create_sheet("Done_PART1")
    ws_done.append(headers)
    for row in done_rows:
        ws_done.append(list(row))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()

    logger.info(
        "Checkpoint: PART1 analysis xlsx %s (Pending_PART1=%s Done_PART1=%s)",
        path.resolve(),
        len(pending_rows),
        len(done_rows),
    )
    logger.info(
        "PART1 tracker compare summary: unmatched_or_missing=%s by_reason=%s",
        sum(missing_or_unmatched.values()),
        dict(missing_or_unmatched),
    )


def data_dir(logger: logging.Logger) -> Path:
    raw = os.environ.get(DATA_DIR_ENV, "")
    base = Path(raw) if raw.strip() else Path(__file__).resolve().parent
    logger.info("Checkpoint: data directory %s", base)
    if not base.is_dir():
        logger.error("HALT: data directory does not exist: %s", base)
        raise SystemExit(1)
    return base


def cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
    return str(val).strip()


def normalize_key_header(s: str) -> str:
    return " ".join(cell_str(s).split())


def normalize_app_number_key(raw: Any) -> str:
    """
    Canonical key for LOS Application Number and COMBINE Application No. joins.
    Handles case/spacing/dash variants that should represent the same application.
    """
    s = cell_str(raw).strip()
    if not s:
        return ""
    s = s.upper()
    # Normalize common Unicode dash variants to ASCII '-'
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    # Remove all spaces so 'HEV - UP - 953' matches 'HEV-UP-953'
    s = "".join(s.split())
    return s


def normalize_close_type_key(raw: Any) -> str:
    return cell_str(raw).strip().lower()


def parse_loan_id(raw: Any, context: str, logger: logging.Logger) -> int | None:
    s = cell_str(raw)
    if not s:
        logger.warning("SKIP: empty value %s", context)
        return None
    try:
        if "." in s:
            return int(Decimal(s))
        return int(s)
    except (ValueError, TypeError, InvalidOperation):
        logger.warning("SKIP: non-integer value=%r %s", raw, context)
        return None


def is_skippable_tenure_emi_placeholder(raw: Any) -> bool:
    """NA / N/A / blank-style markers on COMBINE — row should be skipped, not HALT."""
    s = cell_str(raw).strip().upper().replace(" ", "")
    if not s:
        return True
    return s in ("NA", "N/A", "#N/A", "-", "--", "NIL", "NULL", ".", "…")


def parse_combine_tenure(raw: Any) -> int | None:
    """Tenure for one COMBINE row; returns None if placeholder or unparsable (caller skips row)."""
    if is_skippable_tenure_emi_placeholder(raw):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    s = cell_str(raw).replace(",", "")
    if not s:
        return None
    try:
        if "." in s:
            return int(Decimal(s))
        return int(s)
    except (ValueError, TypeError, InvalidOperation):
        return None


def parse_combine_emi_amount(raw: Any) -> Decimal | None:
    """Monthly EMI for one COMBINE row; None if placeholder or unparsable."""
    if is_skippable_tenure_emi_placeholder(raw):
        return None
    if raw is None or raw == "":
        return None
    if isinstance(raw, int):
        return Decimal(raw)
    if isinstance(raw, float):
        return Decimal(str(raw))
    if isinstance(raw, Decimal):
        return raw
    s = cell_str(raw).replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_combine_total_amount(raw: Any) -> Decimal | None:
    """
    EMI-k Total Amount in COMBINE — never raises. Sheets may put labels like 'End' past active EMIs.
    """
    if raw is None or raw == "":
        return None
    s = cell_str(raw).strip()
    if not s:
        return None
    u = s.upper().replace(" ", "")
    if u in ("END", "NA", "N/A", "#N/A", "-", "--", "NIL", "NULL", "."):
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return Decimal(raw)
    if isinstance(raw, float):
        return Decimal(str(raw))
    if isinstance(raw, Decimal):
        return raw
    s2 = s.replace(",", "")
    try:
        return Decimal(s2)
    except InvalidOperation:
        return None


def parse_decimal(raw: Any, *, context: str, logger: logging.Logger) -> Decimal | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int,)):
        return Decimal(raw)
    if isinstance(raw, float):
        return Decimal(str(raw))
    if isinstance(raw, Decimal):
        return raw
    s = cell_str(raw).replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        logger.error("HALT: bad decimal in %s: %r", context, raw)
        raise RuntimeError(f"HALT: bad decimal in {context}") from None


def is_valid_paid_on(value: Any) -> bool:
    """
    Count a COMBINE 'Paid On' cell as filled when it represents a real calendar payment date.

    Excel often stores dates as serial numbers (float); openpyxl may return float instead of datetime.
    We treat serials in a typical 1980–2080 range as dates. Plain small numbers (e.g. amounts) stay excluded.
    """
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, time):
        return True
    if isinstance(value, (int, float)):
        try:
            v = float(value)
            # Excel 1900 date serials for ~1980–2078 (reject amounts like 5000; accept ~29500–55000+)
            if 25000 <= v <= 80000:
                from_excel(v)
                return True
        except Exception:
            pass
        return False
    s = cell_str(value)
    if not s:
        return False
    # Unparsed numeric-only text: often serial or wrong type; don't count as 'paid' without conversion
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return False
    return True


def find_header_row(
    ws, required: list[str], label: str, logger: logging.Logger, max_scan_rows: int = 80
) -> tuple[int, dict[str, int]]:
    """
    First row (1-based) in which every required header appears as a cell value (normalized whitespace).
    """
    req = [normalize_key_header(x) for x in required]
    req_set = set(req)
    width = _sheet_scan_width(ws)
    for r in range(1, max_scan_rows + 1):
        found: dict[str, int] = {}
        for c in range(1, width + 1):
            v = ws.cell(row=r, column=c).value
            key = normalize_key_header(v)
            if key in req_set and key not in found:
                found[key] = c
        if all(k in found for k in req):
            idx = {orig: found[normalize_key_header(orig)] for orig in required}
            logger.info("Checkpoint: header row for %s is row %s", label, r)
            return r, idx
    seen_sample: list[str] = []
    scan = min(3, ws.max_row or 0)
    w2 = min(width, 40)
    for r in range(1, scan + 1):
        row_vals = []
        for c in range(1, w2 + 1):
            row_vals.append(normalize_key_header(ws.cell(row=r, column=c).value))
        seen_sample.append(",".join(x for x in row_vals if x)[:500])
    msg = (
        f"HALT: required columns not found in {label}. "
        f"Need exact headers {required!r}. Sample rows: {seen_sample!r}"
    )
    logger.error(msg)
    raise RuntimeError(msg)


def find_column_exact(ws, header_row: int, header: str) -> int | None:
    """1-based column index where normalized cell text equals normalized header string."""
    target = normalize_key_header(header)
    for c in range(1, _sheet_scan_width(ws) + 1):
        if normalize_key_header(ws.cell(row=header_row, column=c).value) == target:
            return c
    return None


def _sheet_scan_width(ws) -> int:
    """Safe width for header/body scans (read_only=False sets max_column; otherwise avoid 1-col bug)."""
    mc = ws.max_column
    if mc is not None and mc > 0:
        return int(mc)
    return 512


def load_worksheet(path: Path, sheet_name: str, logger: logging.Logger):
    if not path.is_file():
        logger.error("HALT: file not found: %s", path)
        raise RuntimeError(f"HALT: file not found: {path}")
    logger.info("Checkpoint: opening %s sheet=%s", path.name, sheet_name)
    # read_only=True leaves max_column/max_row None until iteration — breaks header discovery.
    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        logger.error("HALT: sheet %r not in workbook. Sheets: %s", sheet_name, wb.sheetnames)
        wb.close()
        raise RuntimeError(f"HALT: sheet {sheet_name!r} missing")
    return wb[sheet_name], wb


def load_loan_ids(logger: logging.Logger, base: Path) -> list[int]:
    fn, sheet = FILES["db"]
    ws, wb = load_worksheet(base / fn, sheet, logger)
    try:
        hdr_row, col_map = find_header_row(ws, [COL_DB_LOAN_ID], f"{fn}/{sheet}", logger)
        ordered: list[int] = []
        seen: set[int] = set()
        for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
            lid = parse_loan_id(ws.cell(row=r, column=col_map[COL_DB_LOAN_ID]).value, f"row{r}", logger)
            if lid is not None and lid not in seen:
                seen.add(lid)
                ordered.append(lid)
        loan_ids = ordered
    finally:
        wb.close()
    logger.info("Checkpoint: loaded %s distinct loan IDs from %s", len(loan_ids), fn)
    return loan_ids


def load_los_mapping(logger: logging.Logger, base: Path) -> dict[int, str]:
    fn, sheet = FILES["los"]
    ws, wb = load_worksheet(base / fn, sheet, logger)
    try:
        required = [COL_LOS_LOAN_ID, COL_LOS_APP_NUMBER]
        hdr_row, col_map = find_header_row(ws, required, f"{fn}/{sheet}", logger)
        out: dict[int, str] = {}
        excluded: set[int] = set()
        conflict_rows: list[tuple[int, int, str, str]] = []
        for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
            lid = parse_loan_id(
                ws.cell(row=r, column=col_map[COL_LOS_LOAN_ID]).value,
                f"LOS row{r} {COL_LOS_LOAN_ID}",
                logger,
            )
            if lid is None:
                continue
            if lid in excluded:
                continue
            app = cell_str(ws.cell(row=r, column=col_map[COL_LOS_APP_NUMBER]).value)
            if not app:
                logger.warning(
                    "SKIP: LOS row %s missing Application Number for %s=%s",
                    r,
                    COL_LOS_LOAN_ID,
                    lid,
                )
                continue
            if lid in out:
                if out[lid] == app:
                    continue
                prev_app = out[lid]
                logger.warning(
                    "LOS: duplicate application_id=%s with different Application Number: %r vs %r "
                    "(row %s) — excluding this application_id from LOS map (not used downstream)",
                    lid,
                    prev_app,
                    app,
                    r,
                )
                conflict_rows.append((r, lid, prev_app, app))
                excluded.add(lid)
                del out[lid]
                continue
            out[lid] = app
    finally:
        wb.close()
    if conflict_rows:
        bad_ids = sorted({c[1] for c in conflict_rows})
        logger.warning(
            "LOS summary: %s application_id(s) excluded because Application Number conflicted between rows — "
            "these ids are dropped from loan → Application Number mapping: %s",
            len(bad_ids),
            bad_ids,
        )
        for row, lid, a, b in conflict_rows:
            logger.warning(
                "  conflict detail: application_id=%s row=%s Application Number %r vs %r",
                lid,
                row,
                a,
                b,
            )
    logger.info("Checkpoint: LOS map size=%s (from %s)", len(out), fn)
    return out


def load_los_customer_map(logger: logging.Logger, base: Path) -> dict[int, int]:
    fn, sheet = FILES["los"]
    ws, wb = load_worksheet(base / fn, sheet, logger)
    try:
        required = [COL_LOS_LOAN_ID, COL_LOS_CUSTOMER_ID]
        hdr_row, col_map = find_header_row(ws, required, f"{fn}/{sheet} customer map", logger)
        out: dict[int, int] = {}
        for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
            lid = parse_loan_id(
                ws.cell(row=r, column=col_map[COL_LOS_LOAN_ID]).value,
                f"LOS row{r} {COL_LOS_LOAN_ID}",
                logger,
            )
            if lid is None:
                continue
            cust = parse_int_strict(
                ws.cell(row=r, column=col_map[COL_LOS_CUSTOMER_ID]).value,
                f"LOS row{r} {COL_LOS_CUSTOMER_ID}",
                logger,
            )
            if cust is None:
                continue
            if lid not in out:
                out[lid] = cust
    finally:
        wb.close()
    logger.info("Checkpoint: LOS customer map size=%s (from %s)", len(out), fn)
    return out


def scan_combine_grouped_emis(header_row: int, ws, logger: logging.Logger) -> dict[int, dict[str, int]]:
    """
    COMBINE layout: header cells 'EMI - N' then 'Paid On', 'Cash Amount', 'Online Amount',
    'Total Amount', 'Comments' (evaluated on COMBINE-EMI-FOUR-ZONES.xlsx Sheet1 row 1).
    """
    width = _sheet_scan_width(ws)
    out: dict[int, dict[str, int]] = {}
    for c in range(1, width + 1):
        raw = ws.cell(row=header_row, column=c).value
        h = cell_str(raw)
        m = RE_EMI_GROUP_LABEL.match(h)
        if not m:
            continue
        k = int(m.group(1))
        pcol = normalize_key_header(ws.cell(row=header_row, column=c + 1).value)
        ccol = normalize_key_header(ws.cell(row=header_row, column=c + 2).value)
        ocol = normalize_key_header(ws.cell(row=header_row, column=c + 3).value)
        tcol = normalize_key_header(ws.cell(row=header_row, column=c + 4).value)
        cmcol = normalize_key_header(ws.cell(row=header_row, column=c + 5).value)
        if pcol != normalize_key_header(SUB_PAID_ON):
            msg = (
                f"HALT: COMBINE header row {header_row} col {c} {h!r}: "
                f"expected next column {SUB_PAID_ON!r}, found {pcol!r}"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        if ccol != normalize_key_header(SUB_CASH_AMOUNT):
            msg = (
                f"HALT: COMBINE header row {header_row} col {c} {h!r}: "
                f"expected col+2 {SUB_CASH_AMOUNT!r}, found {ccol!r}"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        if ocol != normalize_key_header(SUB_ONLINE_AMOUNT):
            msg = (
                f"HALT: COMBINE header row {header_row} col {c} {h!r}: "
                f"expected col+3 {SUB_ONLINE_AMOUNT!r}, found {ocol!r}"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        if tcol != normalize_key_header(SUB_TOTAL_AMOUNT):
            msg = (
                f"HALT: COMBINE header row {header_row} col {c} {h!r}: "
                f"expected col+4 {SUB_TOTAL_AMOUNT!r}, found {tcol!r}"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        if cmcol != normalize_key_header(SUB_COMMENTS):
            msg = (
                f"HALT: COMBINE header row {header_row} col {c} {h!r}: "
                f"expected col+5 {SUB_COMMENTS!r}, found {cmcol!r}"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        if k in out:
            logger.error("HALT: duplicate EMI block for k=%s in COMBINE header", k)
            raise RuntimeError("HALT: duplicate EMI block")
        out[k] = {"emi_value": c, "paid_on": c + 1, "cash": c + 2, "online": c + 3, "total": c + 4, "comments": c + 5}
    if not out:
        logger.error(
            "HALT: no 'EMI - N' group headers (pattern %r) on row %s",
            RE_EMI_GROUP_LABEL.pattern,
            header_row,
        )
        raise RuntimeError("HALT: no EMI blocks in COMBINE")
    logger.info("Checkpoint: COMBINE grouped EMI columns for k=%s..%s", min(out), max(out))
    return out


def load_combine_rows(
    logger: logging.Logger,
    base: Path,
    app_keys_filter: set[str] | None = None,
) -> dict[str, dict[str, Any]]:
    """
    app_key -> app, tenure, emi, emi_map, emi_value{k}, paid_on{k}, cash{k}, online{k}, total{k}, comments{k}.
    """
    fn, sheet = FILES["combine"]
    path = base / fn
    ws, wb = load_worksheet(path, sheet, logger)
    try:
        base_req = [COL_CT_APP, COL_CT_TENURE, COL_CT_EMI, COL_CT_CLOSE_TYPE]
        hdr_row, col_base = find_header_row(ws, base_req, f"{fn}/{sheet} (base)", logger)
        emi_map = scan_combine_grouped_emis(hdr_row, ws, logger)
        by_app_key: dict[str, dict[str, Any]] = {}
        for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
            app = cell_str(ws.cell(row=r, column=col_base[COL_CT_APP]).value)
            if not app:
                continue
            app_key = normalize_app_number_key(app)
            if app_keys_filter is not None and app_key not in app_keys_filter:
                continue
            tenure_raw = ws.cell(row=r, column=col_base[COL_CT_TENURE]).value
            emi_raw = ws.cell(row=r, column=col_base[COL_CT_EMI]).value
            close_type_raw = ws.cell(row=r, column=col_base[COL_CT_CLOSE_TYPE]).value
            close_type = cell_str(close_type_raw)
            t = parse_combine_tenure(tenure_raw)
            emi = parse_combine_emi_amount(emi_raw)
            if t is None or emi is None:
                logger.warning(
                    "COMBINE row %s: skip row — Tenure/EMI blank, placeholder (e.g. NA), or not numeric "
                    "(Tenure=%r EMI=%r)",
                    r,
                    tenure_raw,
                    emi_raw,
                )
                continue
            for k in range(1, t + 1):
                if k not in emi_map:
                    logger.error(
                        "HALT: COMBINE row %s tenure=%s needs EMI-%s Paid On/Total Amount columns in header",
                        r,
                        t,
                        k,
                    )
                    raise RuntimeError(f"HALT: missing EMI-{k} in header for tenure {t}")
            if app_key in by_app_key:
                logger.warning(
                    "COMBINE row %s: duplicate Application No. key %r (raw=%r) — keeping first row=%s, skipping this row",
                    r,
                    app_key,
                    app,
                    by_app_key[app_key]["row"],
                )
                continue
            emi_value: dict[int, Any] = {}
            paid_on: dict[int, Any] = {}
            cash_amt: dict[int, Any] = {}
            online_amt: dict[int, Any] = {}
            total_amt: dict[int, Any] = {}
            comments: dict[int, Any] = {}
            for k in range(1, t + 1):
                emi_value[k] = ws.cell(row=r, column=emi_map[k]["emi_value"]).value
                paid_on[k] = ws.cell(row=r, column=emi_map[k]["paid_on"]).value
                cash_amt[k] = ws.cell(row=r, column=emi_map[k]["cash"]).value
                online_amt[k] = ws.cell(row=r, column=emi_map[k]["online"]).value
                total_amt[k] = ws.cell(row=r, column=emi_map[k]["total"]).value
                comments[k] = ws.cell(row=r, column=emi_map[k]["comments"]).value
            by_app_key[app_key] = {
                "app": app,
                "row": r,
                "tenure": t,
                "emi": emi,
                "close_type": close_type,
                "emi_map": emi_map,
                "emi_value": emi_value,
                "paid_on": paid_on,
                "cash": cash_amt,
                "online": online_amt,
                "total": total_amt,
                "comments": comments,
            }
    finally:
        wb.close()
    logger.info("Checkpoint: COMBINE indexed %s application keys", len(by_app_key))
    return by_app_key


def parse_int_strict(raw: Any, ctx: str, logger: logging.Logger) -> int | None:
    if raw is None or raw == "":
        logger.error("HALT: missing int for %s", ctx)
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    s = cell_str(raw)
    try:
        return int(Decimal(s))
    except (ValueError, TypeError, InvalidOperation):
        logger.error("HALT: bad int for %s: %r", ctx, raw)
        return None


def build_center_to_manager_map(logger: logging.Logger) -> dict[int, int]:
    center_to_manager: dict[int, int] = {}
    for manager_id, center_ids in CENTER_MANAGER_TO_CENTER_IDS.items():
        for center_id in center_ids:
            if center_id in center_to_manager and center_to_manager[center_id] != manager_id:
                logger.error(
                    "HALT: center_id=%s mapped to multiple manager ids (%s, %s)",
                    center_id,
                    center_to_manager[center_id],
                    manager_id,
                )
                raise RuntimeError("HALT: duplicate center_id mapping to different manager")
            center_to_manager[center_id] = manager_id
    return center_to_manager


def load_dp_sheet(
    logger: logging.Logger, base: Path
) -> tuple[dict[int, dict[int, list[DpEmiRow]]], dict[int, int], int]:
    """
    loan_id -> emi_installment_no -> list of DpEmiRow (usually one; two allowed for PART1 + PART2).

    Rows are typically **pending collections only** — not a full EMI 1..tenure grid; missing EMIs are OK.
    """
    fn, sheet = FILES["dp"]
    required = [
        COL_DP_LOAN_ID,
        COL_DP_TENURE,
        COL_DP_MONTHLY,
        COL_DP_DUE_AMOUNT,
        COL_DP_CENTER_ID,
        COL_DP_COLLECTION_ID,
        COL_DP_EMI_NO,
        COL_DP_STATUS,
        COL_DP_SUBTYPE,
    ]
    ws, wb = load_worksheet(base / fn, sheet, logger)
    try:
        hdr_row, col_map = find_header_row(ws, required, f"{fn}/{sheet}", logger)
        col_p1 = find_column_exact(ws, hdr_row, COL_DP_P1)
        col_p2 = find_column_exact(ws, hdr_row, COL_DP_P2)
        if col_p1 is None:
            logger.info(
                "Checkpoint: optional header %r not found on %s — treating p1 as absent",
                COL_DP_P1,
                fn,
            )
        if col_p2 is None:
            logger.info(
                "Checkpoint: optional header %r not found on %s — treating p2 as absent",
                COL_DP_P2,
                fn,
            )
        by_loan: dict[int, dict[int, list[DpEmiRow]]] = {}
        tenure_by_loan: dict[int, int] = {}
        for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
            lid = parse_loan_id(ws.cell(row=r, column=col_map[COL_DP_LOAN_ID]).value, f"dp row{r}", logger)
            if lid is None:
                continue
            tenure = parse_int_strict(
                ws.cell(row=r, column=col_map[COL_DP_TENURE]).value,
                f"dp row{r} tenure",
                logger,
            )
            monthly = parse_decimal(
                ws.cell(row=r, column=col_map[COL_DP_MONTHLY]).value,
                context=f"dp row{r} monthly_payable",
                logger=logger,
            )
            due_amount = parse_decimal(
                ws.cell(row=r, column=col_map[COL_DP_DUE_AMOUNT]).value,
                context=f"dp row{r} due_amount",
                logger=logger,
            )
            center_id = parse_int_strict(
                ws.cell(row=r, column=col_map[COL_DP_CENTER_ID]).value,
                f"dp row{r} center_id",
                logger,
            )
            cid = parse_int_strict(
                ws.cell(row=r, column=col_map[COL_DP_COLLECTION_ID]).value,
                f"dp row{r} collection_id",
                logger,
            )
            emi_no = parse_int_strict(
                ws.cell(row=r, column=col_map[COL_DP_EMI_NO]).value,
                f"dp row{r} emi_installment_no",
                logger,
            )
            if None in (tenure, monthly, due_amount, center_id, cid, emi_no):
                raise RuntimeError(f"HALT: bad required numeric on dp row {r}")
            st = cell_str(ws.cell(row=r, column=col_map[COL_DP_STATUS]).value)
            if not st:
                logger.error("HALT: empty status dp row %s", r)
                raise RuntimeError("HALT: empty status in dp_sheet")
            sub_raw = ws.cell(row=r, column=col_map[COL_DP_SUBTYPE]).value
            if sub_raw is None or cell_str(sub_raw) == "":
                sub = None
            else:
                sub = cell_str(sub_raw)
            p1c = ws.cell(row=r, column=col_p1).value if col_p1 is not None else None
            p2c = ws.cell(row=r, column=col_p2).value if col_p2 is not None else None
            p1 = (
                parse_decimal(p1c, context=f"dp row{r} p1", logger=logger) if col_p1 is not None else None
            )
            p2 = (
                parse_decimal(p2c, context=f"dp row{r} p2", logger=logger) if col_p2 is not None else None
            )
            if lid not in tenure_by_loan:
                tenure_by_loan[lid] = tenure
            elif tenure_by_loan[lid] != tenure:
                logger.error(
                    "HALT: dp_sheet tenure mismatch for loan_id=%s: %s vs %s (row %s)",
                    lid,
                    tenure_by_loan[lid],
                    tenure,
                    r,
                )
                raise RuntimeError("HALT: inconsistent tenure on dp_sheet for loan")
            em = DpEmiRow(
                collection_id=cid,
                emi_installment_no=emi_no,
                tenure=tenure,
                monthly_payable=monthly,
                due_amount=due_amount,
                center_id=center_id,
                status=st,
                collection_sub_type=sub,
                p1=p1,
                p2=p2,
            )
            bucket = dp_row_subtype_bucket(sub)
            lst = by_loan.setdefault(lid, {}).setdefault(emi_no, [])
            for ex in lst:
                if dp_row_subtype_bucket(ex.collection_sub_type) == bucket:
                    logger.error(
                        "HALT: duplicate dp_sheet row for loan_id=%s emi_installment_no=%s "
                        "with same subtype bucket %r (existing subtype=%r new=%r)",
                        lid,
                        emi_no,
                        bucket,
                        ex.collection_sub_type,
                        sub,
                    )
                    raise RuntimeError("HALT: duplicate dp_sheet EMI row (same subtype bucket)")
            lst.append(em)
    finally:
        wb.close()
    total_dp_rows = sum(sum(len(z) for z in v.values()) for v in by_loan.values())
    logger.info(
        "Checkpoint: dp_sheet loans=%s EMI slot groups=%s total collection rows=%s",
        len(by_loan),
        sum(len(v) for v in by_loan.values()),
        total_dp_rows,
    )
    return by_loan, tenure_by_loan, total_dp_rows


def count_actual_emis_values(
    tenure: int, emi_map: dict[int, dict[str, int]], emi_values: dict[int, Any]
) -> tuple[int, list[tuple[int, Any]]]:
    details: list[tuple[int, Any]] = []
    n = 0
    for k in range(1, tenure + 1):
        if k not in emi_map:
            raise RuntimeError(f"missing EMI map for {k}")
        val = emi_values[k]
        details.append((k, val))
        if is_valid_paid_on(val):
            n += 1
    return n, details


def evaluate_amount_adjustment(
    *,
    tracker_total: Decimal | None,
    due_amount: Decimal,
    logger: logging.Logger,
    ctx: str,
) -> tuple[bool, Decimal, Decimal, Decimal, str | None]:
    """
    Returns (ok, tracker_total, fine_amount, discount_amount, fail_reason).
    Rules depend on AMOUNT_MATCH_MODE:
      - exact: tracker_total must equal due_amount
      - unlimited: always allow diff (fine/discount)
      - capped: allow diff only when <= MAX_ALLOWED_AMOUNT_DIFF
    """
    if tracker_total is None:
        logger.error("HALT: missing Total Amount in tracker %s", ctx)
        return False, Decimal("0"), Decimal("0"), Decimal("0"), "combine_total_missing"
    paid = tracker_total
    mode = cell_str(AMOUNT_MATCH_MODE).strip().lower()
    if mode not in {"exact", "unlimited", "capped"}:
        logger.error("HALT: invalid AMOUNT_MATCH_MODE=%r", AMOUNT_MATCH_MODE)
        return False, paid, Decimal("0"), Decimal("0"), "invalid_amount_match_mode"
    if mode == "exact":
        if paid == due_amount:
            return True, paid, Decimal("0"), Decimal("0"), None
        return False, paid, Decimal("0"), Decimal("0"), "amount_mismatch_exact_required"
    if paid > due_amount:
        extra = paid - due_amount
        if mode == "unlimited" or MAX_ALLOWED_AMOUNT_DIFF is None or extra <= MAX_ALLOWED_AMOUNT_DIFF:
            return True, paid, extra, Decimal("0"), None
        return False, paid, Decimal("0"), Decimal("0"), "extra_amount_exceeds_limit"
    if paid < due_amount:
        diff = due_amount - paid
        if mode == "unlimited" or MAX_ALLOWED_AMOUNT_DIFF is None or diff <= MAX_ALLOWED_AMOUNT_DIFF:
            return True, paid, Decimal("0"), diff, None
        return False, paid, Decimal("0"), Decimal("0"), "shortfall_exceeds_limit"
    return True, paid, Decimal("0"), Decimal("0"), None


def get_db_conn(logger: logging.Logger):
    host = os.environ.get(ENV_DB_HOST, "127.0.0.1")
    name = os.environ.get(ENV_DB_NAME, "myback")
    user = os.environ.get(ENV_DB_USER, "postgres")
    password = os.environ.get(ENV_DB_PASS, "")
    port = os.environ.get(ENV_DB_PORT, "5432")
    logger.info("Checkpoint: connecting PostgreSQL %s:%s db=%s user=%s", host, port, name, user)
    return psycopg2.connect(host=host, port=port, dbname=name, user=user, password=password)


def primary_validate_excel_sources(logger: logging.Logger, base: Path) -> None:
    """
    Fail-fast checks: file exists, sheet exists, non-empty dimensions, required headers,
    COMBINE EMI block pattern, optional dp p1/p2. Logged as PRIMARY for easy grep.
    """
    logger.info("=== PRIMARY CHECK (fail-fast): Excel access + headers for db / LOS / COMBINE / DP ===")

    fn, sh = FILES["db"]
    p = base / fn
    ws, wb = load_worksheet(p, sh, logger)
    try:
        logger.info(
            "PRIMARY: %s / %s — max_row=%s max_column=%s (read_only=False required for dimensions)",
            fn,
            sh,
            ws.max_row,
            ws.max_column,
        )
        find_header_row(ws, [COL_DB_LOAN_ID], f"PRIMARY {fn}/{sh}", logger)
    finally:
        wb.close()

    fn, sh = FILES["los"]
    p = base / fn
    ws, wb = load_worksheet(p, sh, logger)
    try:
        logger.info(
            "PRIMARY: %s / %s — max_row=%s max_column=%s sheets=%s",
            fn,
            sh,
            ws.max_row,
            ws.max_column,
            wb.sheetnames,
        )
        find_header_row(
            ws,
            [COL_LOS_LOAN_ID, COL_LOS_APP_NUMBER, COL_LOS_CUSTOMER_ID],
            f"PRIMARY {fn}/{sh}",
            logger,
        )
        logger.info(
            "PRIMARY: LOS will join DB loan_id to %r (Excel column) + %r",
            COL_LOS_LOAN_ID,
            COL_LOS_APP_NUMBER,
        )
    finally:
        wb.close()

    fn, sh = FILES["combine"]
    p = base / fn
    ws, wb = load_worksheet(p, sh, logger)
    try:
        logger.info("PRIMARY: %s / %s — max_row=%s max_column=%s", fn, sh, ws.max_row, ws.max_column)
        hdr, _ = find_header_row(
            ws,
            [COL_CT_APP, COL_CT_TENURE, COL_CT_EMI, COL_CT_CLOSE_TYPE],
            f"PRIMARY {fn}/{sh} base",
            logger,
        )
        scan_combine_grouped_emis(hdr, ws, logger)
    finally:
        wb.close()

    df, ds = FILES["dp"]
    p = base / df
    ws, wb = load_worksheet(p, ds, logger)
    try:
        logger.info("PRIMARY: DP export %s / %s — max_row=%s max_column=%s", df, ds, ws.max_row, ws.max_column)
        req_dp = [
            COL_DP_LOAN_ID,
            COL_DP_TENURE,
            COL_DP_MONTHLY,
            COL_DP_DUE_AMOUNT,
            COL_DP_CENTER_ID,
            COL_DP_COLLECTION_ID,
            COL_DP_EMI_NO,
            COL_DP_STATUS,
            COL_DP_SUBTYPE,
        ]
        hdr, _ = find_header_row(ws, req_dp, f"PRIMARY {df}/{ds}", logger)
        for opt in (COL_DP_P1, COL_DP_P2):
            c = find_column_exact(ws, hdr, opt)
            logger.info("PRIMARY: optional column %r → %s", opt, f"col {c}" if c else "absent")
        if (df, ds) == (FILES["db"][0], FILES["db"][1]):
            logger.info("PRIMARY: DP rows load from same file/sheet as db_sheet (expected for your layout).")
    finally:
        wb.close()

    logger.info("PRIMARY CHECK passed — loading data.")


def fetch_pending_collection(
    cur, collection_id: int, logger: logging.Logger, ctx: str
) -> tuple[bool, str | None]:
    cur.execute(
        sql.SQL(
            "SELECT status FROM {} WHERE collection_id = %s AND is_active = true"
        ).format(sql.Identifier(TABLE_COLLECTIONS)),
        (collection_id,),
    )
    row = cur.fetchone()
    if row is None:
        logger.error("SKIP: collection_id=%s not in DB %s", collection_id, ctx)
        return False, None
    st = row[0]
    return True, cell_str(st)


def run_backfill(execute: bool, logger: logging.Logger) -> int:
    base = data_dir(logger)
    logger.info("Checkpoint: Script started. Mode=%s", "EXECUTE" if execute else "DRY-RUN")

    primary_validate_excel_sources(logger, base)

    loan_ids = load_loan_ids(logger, base)
    los = load_los_mapping(logger, base)
    los_customer = load_los_customer_map(logger, base)

    los_app_keys = {normalize_app_number_key(v) for v in los.values() if normalize_app_number_key(v)}
    combine_by_app = load_combine_rows(logger, base, los_app_keys)
    dp_by_loan, dp_tenure, dp_total_rows = load_dp_sheet(logger, base)
    center_to_manager = build_center_to_manager_map(logger)

    conn = None
    if execute:
        conn = get_db_conn(logger)
        conn.autocommit = False

    updated = 0
    skipped = 0  # kept: total EMI-level skip events (for compatibility)
    skip_reasons: Counter[str] = Counter()
    mismatches = 0
    mismatch_reasons: Counter[str] = Counter()
    errors = 0
    would_update_collection_ids: set[int] = set()
    updated_report_rows: list[tuple[int, int, str, int, str, str]] = []
    amount_adjustment_rows: list[tuple[int, int, str, int, str, str, str, str, str]] = []
    trans_rows_phase2: list[tuple[Any, ...]] = []
    comment_rows_phase2: list[tuple[Any, ...]] = []
    next_trans_id = 15000
    next_comment_id = 15000
    cannot_update_block: dict[int, dict[str, str]] = {}
    # Tenure gate: COMBINE Tenure == DP tenure only (no_of_actual_emis is informational)
    stats_reached_tenure_compare = 0
    stats_tenure_combine_eq_dp = 0
    stats_tenure_combine_ne_dp = 0

    def append_phase2_rows_for_collection(
        *,
        dp: DpEmiRow,
        loan_id: int,
        emi_no: int,
        manager_id: int,
        cmb: dict[str, Any],
        paid_on_value: Any,
    ) -> None:
        nonlocal next_trans_id, next_comment_id
        created_at = paid_on_to_iso_timestamp(paid_on_value)
        actual_created_at = datetime.now(timezone.utc).isoformat()
        from_customer = los_customer.get(loan_id, "")
        app_no_from_los = cell_str(los.get(loan_id, ""))
        trans_comments = cell_str(cmb.get("app"))
        center_id = dp.center_id
        collection_id = dp.collection_id

        cash_amount = parse_combine_total_amount(cmb["cash"].get(emi_no))
        online_amount = parse_combine_total_amount(cmb["online"].get(emi_no))

        if cash_amount is not None and cash_amount > 0:
            trans_rows_phase2.append(
                (
                    next_trans_id,
                    True,
                    1,
                    center_id,
                    collection_id,
                    True,
                    "",
                    "CREDIT",
                    "RECORD_PAYMENT",
                    "CASH",
                    str(cash_amount),
                    manager_id,
                    0,
                    "ACCEPTED",
                    manager_id,
                    created_at,
                    "",
                    True,
                    0,
                    from_customer,
                    manager_id,
                    manager_id,
                    created_at,
                    trans_comments,
                    "",
                    "",
                    "",
                    actual_created_at,
                )
            )
            next_trans_id += 1

        if online_amount is not None and online_amount > 0:
            trans_rows_phase2.append(
                (
                    next_trans_id,
                    True,
                    1,
                    center_id,
                    collection_id,
                    True,
                    "DUMMYBC2",
                    "CREDIT",
                    "RECORD_PAYMENT",
                    "UPI",
                    str(online_amount),
                    manager_id,
                    0,
                    "ACCEPTED",
                    manager_id,
                    created_at,
                    "",
                    True,
                    0,
                    from_customer,
                    manager_id,
                    manager_id,
                    created_at,
                    trans_comments,
                    "",
                    "",
                    "",
                    actual_created_at,
                )
            )
            next_trans_id += 1

        tracker_comment = cell_str(cmb["comments"].get(emi_no))
        if tracker_comment:
            if app_no_from_los:
                comment_text = f"Application NO. {app_no_from_los} || {tracker_comment}"
            else:
                comment_text = tracker_comment
            comment_rows_phase2.append(
                (
                    next_comment_id,
                    1,
                    collection_id,
                    False,
                    comment_text,
                    False,
                    manager_id,
                    actual_created_at,
                )
            )
            next_comment_id += 1

    try:
        for idx, loan_id in enumerate(loan_ids, 1):
            ctx_base = f"loan_id={loan_id} [{idx}/{len(loan_ids)}]"
            logger.info("Checkpoint: processing %s", ctx_base)

            if loan_id not in los:
                logger.warning("MISMATCH: %s not in LOS mapping — skip", ctx_base)
                mismatches += 1
                mismatch_reasons["loan_not_in_LOS"] += 1
                note_all_pending_collections_for_loan(
                    cannot_update_block, los, dp_by_loan, loan_id, "loan_not_in_LOS"
                )
                continue
            app_key = normalize_app_number_key(los[loan_id])
            if app_key not in combine_by_app:
                logger.warning("MISMATCH: %s no COMBINE row — skip", ctx_base)
                mismatches += 1
                mismatch_reasons["no_COMBINE_for_loan"] += 1
                note_all_pending_collections_for_loan(
                    cannot_update_block, los, dp_by_loan, loan_id, "no_COMBINE_for_loan"
                )
                continue
            if loan_id not in dp_by_loan:
                logger.warning("MISMATCH: %s no dp_sheet rows — skip", ctx_base)
                mismatches += 1
                mismatch_reasons["no_db_sheet_rows_for_loan"] += 1
                continue

            cmb = combine_by_app[app_key]
            close_type_key = normalize_close_type_key(cmb.get("close_type"))
            if close_type_key in SKIP_CLOSE_TYPES:
                logger.warning(
                    "MISMATCH: %s COMBINE CLOSE_TYPE=%r — skip whole loan",
                    ctx_base,
                    cmb.get("close_type"),
                )
                mismatches += 1
                mismatch_reasons["combine_close_type_blocked"] += 1
                note_all_pending_collections_for_loan(
                    cannot_update_block, los, dp_by_loan, loan_id, "combine_close_type_blocked"
                )
                continue
            dp_em = dp_by_loan[loan_id]

            t_tracker = cmb["tenure"]
            t_dp = dp_tenure.get(loan_id)
            if t_dp is None:
                logger.error("HALT: no tenure for %s", ctx_base)
                errors += 1
                note_all_pending_collections_for_loan(
                    cannot_update_block, los, dp_by_loan, loan_id, "dp_tenure_missing"
                )
                continue

            paid_on_vals = cmb["paid_on"]
            emi_vals = cmb["emi_value"]
            n_actual, det = count_actual_emis_values(t_tracker, cmb["emi_map"], emi_vals)
            done_emis_for_loan = 0
            stats_reached_tenure_compare += 1
            if t_tracker != t_dp:
                stats_tenure_combine_ne_dp += 1
                logger.warning(
                    "MISMATCH: %s Tenure COMBINE=%s != DP tenure=%s — processing continues (no skip). "
                    "no_of_actual_emis=%s (counted from EMI-k column date-like values; informational only). "
                    "Detail=%s",
                    ctx_base,
                    t_tracker,
                    t_dp,
                    n_actual,
                    det,
                )
            else:
                stats_tenure_combine_eq_dp += 1
            logger.debug(
                "%s tenure OK: COMBINE=%s DP=%s; no_of_actual_emis=%s "
                "(informational; counted from EMI-k column date-like values)",
                ctx_base,
                t_tracker,
                t_dp,
                n_actual,
            )

            # DP sheet lists only pending collections — not full EMI 1..tenure. Match COMBINE EMI-k only
            # when emi_installment_no exists in dp export for this loan.

            for k in sorted(dp_em.keys()):
                if k < 1 or k > t_tracker:
                    logger.info(
                        "SKIP: %s EMI-%s exists in dp_sheet but is outside COMBINE tenure=%s",
                        ctx_base,
                        k,
                        t_tracker,
                    )
                    skipped += 1
                    skip_reasons["dp_emi_outside_combine_tenure"] += 1
                    for r in dp_em.get(k, []):
                        if cell_str(r.status).upper() == STATUS_PENDING:
                            note_cannot_update_row(
                                cannot_update_block,
                                los,
                                collection_id=r.collection_id,
                                loan_id=loan_id,
                                reason="dp_emi_outside_combine_tenure",
                            )
                    continue
                po = paid_on_vals[k]
                if not is_valid_paid_on(po):
                    rows_k = dp_em.get(k) or []
                    for r in rows_k:
                        if cell_str(r.status).upper() == STATUS_PENDING:
                            note_cannot_update_row(
                                cannot_update_block,
                                los,
                                collection_id=r.collection_id,
                                loan_id=loan_id,
                                reason="combine_paid_on_missing_or_invalid",
                            )
                    continue
                rows_k = dp_em.get(k) or []

                tot_tr = cmb["total"][k]
                tot_dec = parse_combine_total_amount(tot_tr)
                if tot_dec is None:
                    logger.warning(
                        "SKIP: %s EMI-%s Total Amount not usable for payment check (%r) "
                        "(e.g. blank, End/NA, or non-numeric) — skip this EMI",
                        ctx_base,
                        k,
                        tot_tr,
                    )
                    skipped += 1
                    skip_reasons["combine_total_amount_not_usable"] += 1
                    for r in rows_k:
                        if cell_str(r.status).upper() == STATUS_PENDING:
                            note_cannot_update_row(
                                cannot_update_block,
                                los,
                                collection_id=r.collection_id,
                                loan_id=loan_id,
                                reason="combine_total_amount_not_usable",
                            )
                    continue

                pending_rows = [r for r in rows_k if cell_str(r.status).upper() == STATUS_PENDING]
                if not pending_rows:
                    logger.info(
                        "SKIP: %s EMI-%s no row with status=%s (have %s row(s))",
                        ctx_base,
                        k,
                        STATUS_PENDING,
                        len(rows_k),
                    )
                    skipped += 1
                    skip_reasons["no_row_with_status_PENDING"] += 1
                    continue

                updated_any_in_emi = False
                for dp in pending_rows:
                    if is_part_subtype(dp.collection_sub_type):
                        skipped += 1
                        skip_reasons["part_payment_not_handled"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason="part_payment_not_handled",
                        )
                        logger.info(
                            "SKIP: %s EMI-%s collection_id=%s subtype=%r part payment is not handled",
                            ctx_base,
                            k,
                            dp.collection_id,
                            dp.collection_sub_type,
                        )
                        continue
                    ok_pay, paid_sum, fine_amount, discount_amount, fail_reason = evaluate_amount_adjustment(
                        tracker_total=tot_dec,
                        due_amount=dp.due_amount,
                        logger=logger,
                        ctx=f"{ctx_base} EMI-{k} cid={dp.collection_id}",
                    )
                    if not ok_pay:
                        logger.warning(
                            "SKIP: %s EMI-%s amount rule failed paid=%s due_amount=%s "
                            "collection_id=%s subtype=%r reason=%s",
                            ctx_base,
                            k,
                            paid_sum,
                            dp.due_amount,
                            dp.collection_id,
                            dp.collection_sub_type,
                            fail_reason,
                        )
                        skipped += 1
                        skip_reasons[fail_reason or "amount_rule_failed"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason=fail_reason or "amount_rule_failed",
                        )
                        continue
                    manager_id = center_to_manager.get(dp.center_id)
                    if manager_id is None:
                        logger.warning(
                            "SKIP: %s EMI-%s collection_id=%s center_id=%s has no manager mapping",
                            ctx_base,
                            k,
                            dp.collection_id,
                            dp.center_id,
                        )
                        skipped += 1
                        skip_reasons["center_id_manager_mapping_missing"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason="center_id_manager_mapping_missing",
                        )
                        continue

                    if conn is None:
                        logger.info(
                            "[DRY-RUN] Would UPDATE collection_id=%s → DONE fine=%s discount=%s created_by=%s collector_id=%s (%s EMI-%s, subtype=%r)",
                            dp.collection_id,
                            fine_amount,
                            discount_amount,
                            manager_id,
                            manager_id,
                            ctx_base,
                            k,
                            dp.collection_sub_type,
                        )
                        updated += 1
                        would_update_collection_ids.add(dp.collection_id)
                        updated_report_rows.append(
                            (
                                dp.collection_id,
                                loan_id,
                                cell_str(los.get(loan_id, "")),
                                k,
                                cell_str(dp.collection_sub_type),
                                str(dp.due_amount),
                            )
                        )
                        append_phase2_rows_for_collection(
                            dp=dp,
                            loan_id=loan_id,
                            emi_no=k,
                            manager_id=manager_id,
                            cmb=cmb,
                            paid_on_value=po,
                        )
                        if fine_amount > 0 or discount_amount > 0:
                            amount_adjustment_rows.append(
                                (
                                    dp.collection_id,
                                    loan_id,
                                    cell_str(los.get(loan_id, "")),
                                    k,
                                    cell_str(dp.collection_sub_type),
                                    str(paid_sum),
                                    str(dp.due_amount),
                                    str(fine_amount),
                                    str(discount_amount),
                                )
                            )
                        updated_any_in_emi = True
                        continue

                    cur = conn.cursor()
                    exists, db_status = fetch_pending_collection(
                        cur, dp.collection_id, logger, f"{ctx_base} EMI-{k} cid={dp.collection_id}"
                    )
                    if not exists:
                        skipped += 1
                        skip_reasons["DB_collection_row_not_found_or_inactive"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason="DB_collection_row_not_found_or_inactive",
                        )
                        cur.close()
                        continue
                    if db_status and db_status.upper() != STATUS_PENDING:
                        logger.info("SKIP: DB status=%s for collection_id=%s", db_status, dp.collection_id)
                        skipped += 1
                        skip_reasons["DB_status_not_PENDING"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason="DB_status_not_PENDING",
                        )
                        cur.close()
                        continue

                    cur.execute(
                        sql.SQL(
                            "UPDATE {} SET status = %s, created_by = %s, collector_id = %s, fine_amount = %s, discount_amount = %s "
                            "WHERE collection_id = %s AND is_active = true AND UPPER(status) = %s"
                        ).format(sql.Identifier(TABLE_COLLECTIONS)),
                        (
                            STATUS_DONE,
                            manager_id,
                            manager_id,
                            fine_amount,
                            discount_amount,
                            dp.collection_id,
                            STATUS_PENDING,
                        ),
                    )
                    if cur.rowcount != 1:
                        logger.warning(
                            "SKIP: UPDATE rowcount=%s for collection_id=%s (expected 1)",
                            cur.rowcount,
                            dp.collection_id,
                        )
                        skipped += 1
                        skip_reasons["UPDATE_rowcount_not_1"] += 1
                        note_cannot_update_row(
                            cannot_update_block,
                            los,
                            collection_id=dp.collection_id,
                            loan_id=loan_id,
                            reason="UPDATE_rowcount_not_1",
                        )
                    else:
                        logger.info(
                            "OK: updated collection_id=%s %s EMI-%s fine=%s discount=%s (subtype=%r)",
                            dp.collection_id,
                            ctx_base,
                            k,
                            fine_amount,
                            discount_amount,
                            dp.collection_sub_type,
                        )
                        updated += 1
                        would_update_collection_ids.add(dp.collection_id)
                        updated_report_rows.append(
                            (
                                dp.collection_id,
                                loan_id,
                                cell_str(los.get(loan_id, "")),
                                k,
                                cell_str(dp.collection_sub_type),
                                str(dp.due_amount),
                            )
                        )
                        append_phase2_rows_for_collection(
                            dp=dp,
                            loan_id=loan_id,
                            emi_no=k,
                            manager_id=manager_id,
                            cmb=cmb,
                            paid_on_value=po,
                        )
                        if fine_amount > 0 or discount_amount > 0:
                            amount_adjustment_rows.append(
                                (
                                    dp.collection_id,
                                    loan_id,
                                    cell_str(los.get(loan_id, "")),
                                    k,
                                    cell_str(dp.collection_sub_type),
                                    str(paid_sum),
                                    str(dp.due_amount),
                                    str(fine_amount),
                                    str(discount_amount),
                                )
                            )
                        updated_any_in_emi = True
                    cur.close()
                if updated_any_in_emi:
                    done_emis_for_loan += 1
            logger.info(
                "Checkpoint: %s done_emis=%s (EMI-k with Paid On + Total Amount/payment condition passed)",
                ctx_base,
                done_emis_for_loan,
            )
        if conn:
            conn.commit()
            logger.info("Checkpoint: database commit done.")
    except Exception:
        if conn:
            conn.rollback()
            logger.error("Checkpoint: database rollback after error.")
        raise
    finally:
        if conn:
            conn.close()

    for lid, emap in dp_by_loan.items():
        for rows in emap.values():
            for r in rows:
                if cell_str(r.status).upper() != STATUS_PENDING:
                    continue
                cid = r.collection_id
                if cid in would_update_collection_ids:
                    continue
                if cid not in cannot_update_block:
                    note_cannot_update_row(
                        cannot_update_block,
                        los,
                        collection_id=cid,
                        loan_id=lid,
                        reason="no_eligible_path_COMBINE_paid_on_missing_or_invalid_for_this_emi",
                    )

    report_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = base / f"cannot_update_collections_{report_ts}.csv"
    write_cannot_update_collections_csv(
        report_path, cannot_update_block, would_update_collection_ids, logger
    )
    all_report_path = base / f"all_collection_ids_{report_ts}.csv"
    write_all_collection_ids_csv(
        all_report_path,
        dp_by_loan,
        los,
        cannot_update_block,
        would_update_collection_ids,
        logger,
    )
    updated_xlsx_path = base / f"updated_collections_{report_ts}.xlsx"
    write_updated_collections_xlsx(updated_xlsx_path, updated_report_rows, logger)
    amount_xlsx_path = base / f"amount_adjustments_{report_ts}.xlsx"
    write_amount_adjustments_xlsx(amount_xlsx_path, amount_adjustment_rows, logger)
    part1_xlsx_path = base / f"part1_tracker_analysis_{report_ts}.xlsx"
    write_part1_tracker_analysis_xlsx(part1_xlsx_path, dp_by_loan, los, combine_by_app, logger)
    phase2_dir = base / "produced_sheets_phase_2"
    logger.info("PHASE 2: preparing output sheets in %s", phase2_dir.resolve())
    phase2_trans_path = phase2_dir / f"collection_trans_{report_ts}.csv"
    write_phase2_collection_trans_csv(phase2_trans_path, trans_rows_phase2, logger)
    phase2_comments_path = phase2_dir / f"collection_comments_{report_ts}.csv"
    write_phase2_collection_comments_csv(phase2_comments_path, comment_rows_phase2, logger)
    phase2_mode_counts: Counter[str] = Counter()
    for tr in trans_rows_phase2:
        # collection_trans column order: ... mode at index 9
        mode = cell_str(tr[9]).upper()
        if mode:
            phase2_mode_counts[mode] += 1
    phase2_trans_total = len(trans_rows_phase2)
    phase2_comments_total = len(comment_rows_phase2)
    phase2_total_rows = phase2_trans_total + phase2_comments_total
    due_sum_updated = sum(Decimal(row[5]) for row in updated_report_rows)
    trans_amount_sum = sum(Decimal(str(row[10])) for row in trans_rows_phase2)
    due_vs_trans_match = due_sum_updated == trans_amount_sum

    all_pending_collection_ids: set[int] = set()
    target_pending_collection_ids: set[int] = set()
    for emap in dp_by_loan.values():
        for rows in emap.values():
            for r in rows:
                if cell_str(r.status).upper() != STATUS_PENDING:
                    continue
                all_pending_collection_ids.add(r.collection_id)
                if not is_part_subtype(r.collection_sub_type):
                    target_pending_collection_ids.add(r.collection_id)
    updated_target_collection_ids = would_update_collection_ids & target_pending_collection_ids
    skipped_target_pending_collections = max(0, len(target_pending_collection_ids) - len(updated_target_collection_ids))
    skipped_all_pending_collections = max(0, len(all_pending_collection_ids) - len(would_update_collection_ids))

    dp_file = str(FILES["dp"])
    logger.info("========== RUN SUMMARY ==========")
    logger.info("Source: %s", dp_file)
    logger.info(
        "Mode: %s",
        "EXECUTE (DB updated)" if execute else "DRY-RUN (no DB writes)",
    )
    logger.info(
        "Input scope: total export rows=%s | pending(all)=%s | pending(target, non-part)=%s",
        dp_total_rows,
        len(all_pending_collection_ids),
        len(target_pending_collection_ids),
    )
    logger.info(
        "Result: %s=%s | unique collection_id=%s",
        "updated" if execute else "would_update",
        updated,
        len(would_update_collection_ids),
    )
    logger.info(
        "Pending skipped: target=%s | all_pending=%s",
        skipped_target_pending_collections,
        skipped_all_pending_collections,
    )
    export_rows_not_updated = max(0, dp_total_rows - len(would_update_collection_ids))
    logger.info(
        "Export rows not updated (approx): %s "
        "(= %s - %s unique collection_id %s)",
        export_rows_not_updated,
        dp_total_rows,
        len(would_update_collection_ids),
        "updated" if execute else "would_update",
    )
    logger.info(
        "Loan-level skips: %s | reasons=%s",
        mismatches,
        dict(mismatch_reasons),
    )
    logger.info(
        "Collection/EMI skip decisions: total=%s | reasons=%s",
        skipped,
        dict(skip_reasons),
    )
    logger.info(
        "Part payment skipped (policy): %s",
        skip_reasons.get("part_payment_not_handled", 0),
    )
    logger.info(
        "PHASE 2 summary: collection_trans=%s | by_mode=%s | collection_comments=%s | total_generated_rows=%s",
        phase2_trans_total,
        dict(phase2_mode_counts),
        phase2_comments_total,
        phase2_total_rows,
    )
    logger.info("---------- CHECKS ----------")
    logger.info(
        "Check due_vs_trans_amount: due_sum_from_updated_collections=%s | trans_amount_sum=%s | match=%s",
        due_sum_updated,
        trans_amount_sum,
        due_vs_trans_match,
    )
    logger.info(
        "Errors (HALT paths): %s",
        errors,
    )
    logger.info(
        "Summary tenure: COMBINE Tenure == DP tenure matched=%s not_matched=%s "
        "(among %s loans that reached tenure comparison; LOS+COMBINE+dp must already be OK)",
        stats_tenure_combine_eq_dp,
        stats_tenure_combine_ne_dp,
        stats_reached_tenure_compare,
    )
    logger.info(
        "Note: no_of_actual_emis counts COMBINE EMI-k 'Paid On' cells with a date (including Excel serials "
        "in ~25k–80k). It is logged for audit only and does not block updates."
    )
    logger.info("=================================")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Backfill collections from Excel + dp_sheet.")
    p.add_argument(
        "--execute",
        action="store_true",
        help="Apply updates (default: dry-run).",
    )
    args = p.parse_args()
    logger = setup_logging()
    load_env_file(logger, Path(__file__).resolve().parent / ".env")
    try:
        return run_backfill(execute=args.execute, logger=logger)
    except RuntimeError as e:
        logger.error("FAILED: %s", e)
        return 1
    except Exception:
        logging.getLogger("backfill_collection").exception("FAILED: unexpected error")
        return 1


if __name__ == "__main__":
    sys.exit(main())
