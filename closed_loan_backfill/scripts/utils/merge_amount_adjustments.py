#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def header_map(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        key = cell_str(ws.cell(row=1, column=c).value)
        if key and key not in out:
            out[key] = c
    return out


def first_sheet(path: Path):
    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return wb, ws


def build_remarks_by_collection_id(refine_ws) -> dict[str, str]:
    hm = header_map(refine_ws)
    if "collection_id" not in hm:
        raise RuntimeError("Refine sheet missing required header: collection_id")
    if "Remarks" not in hm:
        raise RuntimeError("Refine sheet missing required header: Remarks")

    cid_col = hm["collection_id"]
    remarks_col = hm["Remarks"]
    out: dict[str, str] = {}
    for r in range(2, (refine_ws.max_row or 1) + 1):
        cid = cell_str(refine_ws.cell(row=r, column=cid_col).value)
        if not cid:
            continue
        remarks = cell_str(refine_ws.cell(row=r, column=remarks_col).value)
        # Keep the first non-empty remark if duplicates appear
        if cid not in out:
            out[cid] = remarks
        elif not out[cid] and remarks:
            out[cid] = remarks
    return out


def merge(base_ws, remarks_by_cid: dict[str, str]) -> Workbook:
    hm = header_map(base_ws)
    if "collection_id" not in hm:
        raise RuntimeError("Base sheet missing required header: collection_id")
    cid_col = hm["collection_id"]

    base_headers: list[str] = [
        cell_str(base_ws.cell(row=1, column=c).value) for c in range(1, (base_ws.max_column or 0) + 1)
    ]
    # ensure Remarks exists at end
    if "Remarks" in base_headers:
        remarks_idx = base_headers.index("Remarks")
    else:
        base_headers.append("Remarks")
        remarks_idx = len(base_headers) - 1

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "amound_adjuatment_final"
    out_ws.append(base_headers)

    for r in range(2, (base_ws.max_row or 1) + 1):
        row_vals = [base_ws.cell(row=r, column=c).value for c in range(1, (base_ws.max_column or 0) + 1)]
        if len(row_vals) < len(base_headers):
            row_vals.extend([""] * (len(base_headers) - len(row_vals)))

        cid = cell_str(base_ws.cell(row=r, column=cid_col).value)
        remarks = remarks_by_cid.get(cid, "")
        row_vals[remarks_idx] = remarks
        out_ws.append(row_vals)

    return out_wb


def main() -> int:
    p = argparse.ArgumentParser(description="Merge amount adjustments with remarks by collection_id.")
    p.add_argument(
        "--refine",
        default="amount_adjustments_refine.xlsx",
        help="Path to refine workbook (contains Remarks).",
    )
    p.add_argument(
        "--base",
        default="amount_adjustments_20260423_141515.xlsx",
        help="Path to base amount_adjustments workbook.",
    )
    p.add_argument(
        "--out",
        default="amound_adjuatment_final.xlsx",
        help="Output workbook path.",
    )
    args = p.parse_args()

    refine_path = Path(args.refine).resolve()
    base_path = Path(args.base).resolve()
    out_path = Path(args.out).resolve()

    refine_wb, refine_ws = first_sheet(refine_path)
    base_wb, base_ws = first_sheet(base_path)
    try:
        remarks_map = build_remarks_by_collection_id(refine_ws)
        out_wb = merge(base_ws, remarks_map)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_wb.save(str(out_path))
        out_wb.close()
        print(f"Created: {out_path}")
    finally:
        refine_wb.close()
        base_wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

