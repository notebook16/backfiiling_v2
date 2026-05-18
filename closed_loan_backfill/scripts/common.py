#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import os
import signal
import sys
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from psycopg2.pool import ThreadedConnectionPool


DEFAULT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SHEETS = DEFAULT_ROOT / "source_sheets"
DEFAULT_CHECKPOINTS = DEFAULT_ROOT / "state"


@dataclass
class Paths:
    root: Path = DEFAULT_ROOT
    source_sheets: Path = DEFAULT_SHEETS
    generated_sheets: Path = field(default_factory=lambda: DEFAULT_ROOT / "generated_sheets")
    generated_collections: Path = field(default_factory=lambda: DEFAULT_ROOT / "generated_collections")
    generated_collection_trans: Path = field(default_factory=lambda: DEFAULT_ROOT / "generated_collection_trans")
    # DB import staging: collections / collection_trans / collection_comments per script.
    generated_db_sheets: Path = field(default_factory=lambda: DEFAULT_ROOT / "generated_DB_sheets")
    logs: Path = field(default_factory=lambda: DEFAULT_ROOT / "logs")
    checkpoints: Path = DEFAULT_CHECKPOINTS


@dataclass
class RuntimeConfig:
    execute: bool = False
    dry_run: bool = True
    resume: bool = False
    max_workers: int = 4
    batch_size: int = 100
    retries: int = 3
    retry_sleep_seconds: float = 1.0


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def ts_label() -> str:
    return utc_now().strftime("%Y%m%d_%H%M%S")


def prompt_db_import(
    logger: logging.Logger,
    *,
    stage_label: str,
    files: list[tuple[str, Path]],
    prompt_text: str | None = None,
) -> bool:
    """Ask whether to apply staged DB work. Returns False on EOF, empty answer, or n."""
    if not files:
        return False
    question = prompt_text or "Import these generated files into the database? (y/N): "
    logger.info("--- staged for database import (%s) ---", stage_label)
    for label, path in files:
        logger.info("  %s: %s", label, path)
    try:
        ans = input(question).strip().lower()
    except EOFError:
        ans = ""
    if ans in {"y", "yes"}:
        return True
    logger.info("database import skipped (answer was not y/yes)")
    return False


def setup_logger(stage: str, paths: Paths) -> logging.Logger:
    paths.logs.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger(f"closed_loan_backfill.{stage}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s")
    fh = logging.FileHandler(paths.logs / f"{stage}_{ts_label()}.log", encoding="utf-8")
    sh = logging.StreamHandler(sys.stdout)
    fh.setFormatter(fmt)
    sh.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def load_env_file(path: Path, logger: logging.Logger | None = None) -> int:
    if not path.is_file():
        if logger:
            logger.info("env file not found at %s; continuing with current environment", path)
        return 0

    loaded = 0
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ[key] = value
        loaded += 1

    if logger:
        logger.info("loaded %s env vars from %s", loaded, path)
    return loaded


def parse_common_args(description: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Legacy: marks non-dry-run runtime; script 3/4 use interactive import after staging unless --dry-run",
    )
    parser.add_argument("--dry-run", action="store_true", help="Force dry-run mode")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint")
    parser.add_argument("--max-workers", type=int, default=4)
    parser.add_argument("--batch-size", type=int, default=100)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--retry-sleep-seconds", type=float, default=1.0)
    parser.add_argument("--root-dir", type=str, default=str(DEFAULT_ROOT))
    parser.add_argument("--source-dir", type=str, default=str(DEFAULT_SHEETS))
    return parser


def runtime_from_args(args: argparse.Namespace) -> tuple[RuntimeConfig, Paths]:
    root = Path(args.root_dir).resolve()
    paths = Paths(root=root, source_sheets=Path(args.source_dir).resolve())
    paths.generated_sheets = root / "generated_sheets"
    paths.generated_db_sheets = root / "generated_DB_sheets"
    paths.generated_collections = root / "generated_collections"
    paths.generated_collection_trans = root / "generated_collection_trans"
    paths.logs = root / "logs"
    paths.checkpoints = root / "state"
    runtime = RuntimeConfig(
        execute=bool(args.execute),
        dry_run=bool(args.dry_run) or not bool(args.execute),
        resume=bool(args.resume),
        max_workers=max(1, int(args.max_workers)),
        batch_size=max(1, int(args.batch_size)),
        retries=max(1, int(args.retries)),
        retry_sleep_seconds=max(0.0, float(args.retry_sleep_seconds)),
    )
    ensure_structure(paths)
    return runtime, paths


def ensure_structure(paths: Paths) -> None:
    paths.root.mkdir(parents=True, exist_ok=True)
    paths.generated_sheets.mkdir(parents=True, exist_ok=True)
    for i in range(1, 6):
        (paths.generated_sheets / f"script_{i}").mkdir(parents=True, exist_ok=True)
    paths.generated_collections.mkdir(parents=True, exist_ok=True)
    paths.generated_collection_trans.mkdir(parents=True, exist_ok=True)
    paths.generated_db_sheets.mkdir(parents=True, exist_ok=True)
    for rel in (
        "script_3/collection_trans",
        "script_3/collection_comments",
        "script_3/collections",
        "script_4/collections",
        "script_4/collection_trans",
    ):
        (paths.generated_db_sheets / rel).mkdir(parents=True, exist_ok=True)
    paths.logs.mkdir(parents=True, exist_ok=True)
    paths.checkpoints.mkdir(parents=True, exist_ok=True)


def read_excel(sheet_path: Path, sheet_name: str | int = 0) -> pd.DataFrame:
    return pd.read_excel(sheet_path, sheet_name=sheet_name)


def read_excel_headers(sheet_path: Path, sheet_name: str | int = 0) -> list[str]:
    wb = load_workbook(sheet_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        return [str(c).strip() if c is not None else "" for c in first]
    finally:
        wb.close()


def to_records(df: pd.DataFrame) -> list[dict[str, Any]]:
    clean = df.where(pd.notna(df), None)
    return clean.to_dict(orient="records")


def write_audit_xlsx(base_dir: Path, prefix: str, buckets: dict[str, list[dict[str, Any]]]) -> dict[str, Path]:
    out_paths: dict[str, Path] = {}
    run_ts = ts_label()
    for bucket, rows in buckets.items():
        output = base_dir / f"{prefix}_{bucket}_{run_ts}.xlsx"
        pd.DataFrame(rows).to_excel(output, index=False)
        out_paths[bucket] = output
    return out_paths


def count_reasons(
    rows: Iterable[dict[str, Any]],
    *field_names: str,
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        reason = ""
        for field_name in field_names:
            value = row.get(field_name)
            if value is not None and str(value).strip():
                reason = str(value).strip()
                break
        if not reason:
            reason = "unknown"
        counts[reason] = counts.get(reason, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))


def log_stage_summary(
    logger: logging.Logger,
    stage: str,
    *,
    loaded: int | None = None,
    buckets: dict[str, list[dict[str, Any]]],
    reason_fields_by_bucket: dict[str, tuple[str, ...]] | None = None,
) -> None:
    logger.info("=== %s summary ===", stage)
    if loaded is not None:
        logger.info("loaded records: %s", loaded)
    for bucket_name, rows in buckets.items():
        logger.info("%s count: %s", bucket_name, len(rows))
        fields = (reason_fields_by_bucket or {}).get(bucket_name, ())
        if fields and rows:
            logger.info("%s by reason: %s", bucket_name, count_reasons(rows, *fields))


class CheckpointStore:
    def __init__(self, stage: str, paths: Paths):
        self._path = paths.checkpoints / f"{stage}.json"
        self._lock = threading.Lock()
        if not self._path.exists():
            self.save({"completed_keys": [], "metadata": {}})

    def load(self) -> dict[str, Any]:
        return json.loads(self._path.read_text(encoding="utf-8"))

    def save(self, payload: dict[str, Any]) -> None:
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def completed(self) -> set[str]:
        data = self.load()
        return set(data.get("completed_keys", []))

    def mark_completed(self, key: str) -> None:
        with self._lock:
            data = self.load()
            completed = set(data.get("completed_keys", []))
            completed.add(str(key))
            data["completed_keys"] = sorted(completed)
            self.save(data)


def retry(logger: logging.Logger, tries: int, sleep_seconds: float) -> Callable[[Callable[..., Any]], Callable[..., Any]]:
    def deco(func: Callable[..., Any]) -> Callable[..., Any]:
        def wrapped(*args: Any, **kwargs: Any) -> Any:
            last: Exception | None = None
            for attempt in range(1, tries + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as exc:  # noqa: BLE001
                    last = exc
                    logger.warning("attempt=%s failed for %s: %s", attempt, func.__name__, exc)
                    if attempt < tries:
                        time.sleep(sleep_seconds)
            raise RuntimeError(f"{func.__name__} failed after {tries} attempts") from last

        return wrapped

    return deco


class DbClient:
    def __init__(self, logger: logging.Logger, minconn: int = 1, maxconn: int = 6):
        self.logger = logger
        self.pool = ThreadedConnectionPool(
            minconn=minconn,
            maxconn=maxconn,
            host=os.environ.get("DB_POSTGRES_URL", "127.0.0.1"),
            port=os.environ.get("DB_POSTGRES_PORT", "5432"),
            dbname=os.environ.get("DB_POSTGRES_DBNAME", "myback"),
            user=os.environ.get("DB_POSTGRES_USERNAME", "postgres"),
            password=os.environ.get("DB_POSTGRES_PASS", ""),
        )

    @contextmanager
    def conn(self):
        conn = self.pool.getconn()
        conn.autocommit = False
        try:
            yield conn
            conn.commit()
        except Exception:  # noqa: BLE001
            conn.rollback()
            raise
        finally:
            self.pool.putconn(conn)

    @contextmanager
    def cursor(self):
        with self.conn() as conn:
            with conn.cursor() as cur:
                yield cur

    def close(self) -> None:
        self.pool.closeall()


def graceful_shutdown(logger: logging.Logger) -> threading.Event:
    stop_event = threading.Event()

    def handler(signum: int, frame: Any) -> None:  # noqa: ARG001
        logger.warning("received signal=%s, stopping gracefully", signum)
        stop_event.set()

    signal.signal(signal.SIGINT, handler)
    signal.signal(signal.SIGTERM, handler)
    return stop_event


def mandatory_failure_fields(
    loan_id: Any,
    application_id: Any,
    reason: str,
    stage: str,
) -> dict[str, Any]:
    return {
        "loan_id": loan_id,
        "application_id": application_id,
        "failure_reason": reason,
        "stage_name": stage,
        "timestamp": utc_now().isoformat(),
    }
